"""Monthly payroll Excel (.xlsx) builder — the "Payrun Employee Salary statement".

Pure and deterministic (no I/O beyond an in-memory buffer). Given the per-employee
rows for a month, produce a single flat sheet in the 40-column Indian payroll
register layout (one row per employee) and return the raw bytes.

Money is written as whole-currency **numbers** so totals sum in Excel; day counts
(Base Days / Loss Of Pay / Effective Paid Days) use a fixed 30-day base per the
register convention. Decoupled from the ORM, like `payslip_pdf`.
"""

from __future__ import annotations

from dataclasses import dataclass, fields
from datetime import date
from io import BytesIO
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Column kinds drive the cell number format: text / money (whole ₹) / days (2dp).
_TEXT, _MONEY, _DAYS = "text", "money", "days"

# (header, field-name, kind) in exact register order — matches the reference sheet.
_COLUMNS: tuple[tuple[str, str, str], ...] = (
    ("Period", "period", _TEXT),
    ("Payroll Type", "payroll_type", _TEXT),
    ("Employee No", "employee_no", _TEXT),
    ("Employee Name", "employee_name", _TEXT),
    ("Department", "department", _TEXT),
    ("Designation", "designation", _TEXT),
    ("Work Location", "work_location", _TEXT),
    ("Date of Joining", "date_of_joining", _TEXT),
    ("Date of Birth", "date_of_birth", _TEXT),
    ("Last Working Day", "last_working_day", _TEXT),
    ("Payment Mode", "payment_mode", _TEXT),
    ("Account Holder Name", "account_holder", _TEXT),
    ("Bank Name", "bank_name", _TEXT),
    ("Account Number", "account_number", _TEXT),
    ("IFSC", "ifsc", _TEXT),
    ("CTC Amount(Per Annum)", "ctc_annual", _MONEY),
    ("Gross Amount(Per Annum)", "gross_annual", _MONEY),
    ("Base Days", "base_days", _DAYS),
    ("Loss Of Pay", "loss_of_pay", _DAYS),
    ("Effective Paid Days", "effective_paid_days", _DAYS),
    ("Basic", "basic", _MONEY),
    ("House Rent Allowance", "hra", _MONEY),
    ("Fixed Allowance", "fixed_allowance", _MONEY),
    ("Reimbursement", "reimbursement", _MONEY),
    ("Total Reimbursements", "total_reimbursements", _MONEY),
    ("Fixed Monthly Earnings", "fixed_monthly_earnings", _MONEY),
    ("Fixed Monthly Costs (Earnings + Employer Contributions)", "fixed_monthly_costs", _MONEY),
    ("Total Earnings", "total_earnings", _MONEY),
    ("EPF Contribution", "epf_employee", _MONEY),
    ("EPF Contribution Employer", "epf_employer", _MONEY),
    ("EPS Contribution Employer", "eps_employer", _MONEY),
    ("Employer EDLI Contribution Employer", "edli_employer", _MONEY),
    ("Employer EPF Admin Charges Employer", "epf_admin_employer", _MONEY),
    ("Total Employer Contributions", "total_employer_contributions", _MONEY),
    ("Income Tax", "income_tax", _MONEY),
    ("Maharashtra Professional Tax", "professional_tax", _MONEY),
    ("Total Deductions", "total_deductions", _MONEY),
    ("Gross Pay", "gross_pay", _MONEY),
    ("Net Pay", "net_pay", _MONEY),
    ("Business Expense Reimbursements", "business_expense_reimbursements", _MONEY),
)

_HEADER_FILL = PatternFill("solid", fgColor="5A48E0")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def fmt_register_date(value: date | None) -> str:
    """Register date format (dd-mm-yyyy); empty string when unset."""
    return value.strftime("%d-%m-%Y") if value is not None else ""


@dataclass(frozen=True)
class PayrollExportRow:
    """One employee's line in the register. Money fields are whole currency units;
    day fields use a 30-day base. Dates are pre-formatted to strings so the sheet
    reads uniformly regardless of the viewer's locale."""

    period: str
    payroll_type: str
    employee_no: str
    employee_name: str
    department: str
    designation: str
    work_location: str
    date_of_joining: str
    date_of_birth: str
    last_working_day: str
    payment_mode: str
    account_holder: str
    bank_name: str
    account_number: str
    ifsc: str
    ctc_annual: int
    gross_annual: int
    base_days: float
    loss_of_pay: float
    effective_paid_days: float
    basic: int
    hra: int
    fixed_allowance: int
    reimbursement: int
    total_reimbursements: int
    fixed_monthly_earnings: int
    fixed_monthly_costs: int
    total_earnings: int
    epf_employee: int
    epf_employer: int
    eps_employer: int
    edli_employer: int
    epf_admin_employer: int
    total_employer_contributions: int
    income_tax: int
    professional_tax: int
    total_deductions: int
    gross_pay: int
    net_pay: int
    business_expense_reimbursements: int

    def as_cells(self) -> list[object]:
        return [getattr(self, name) for _header, name, _kind in _COLUMNS]


# Fail loudly if the dataclass and the column spec ever drift apart.
assert {name for _h, name, _k in _COLUMNS} == {f.name for f in fields(PayrollExportRow)}


def _autofit(ws: Worksheet, rows: list[PayrollExportRow]) -> None:
    for idx, (header, name, _kind) in enumerate(_COLUMNS, start=1):
        widest = len(header)
        for row in rows:
            widest = max(widest, len(str(getattr(row, name))))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(widest + 2, 10), 44)


def build_payroll_xlsx(rows: list[PayrollExportRow], *, month_label: str, currency: str) -> bytes:
    """Render the month's payroll rows to .xlsx bytes in the register layout."""
    wb = Workbook()
    ws = cast(Worksheet, wb.active)  # a fresh Workbook always has an active sheet
    ws.title = "Payrun Employee Salary statemen"  # 31-char cap (matches the reference)

    ws.append([header for header, _name, _kind in _COLUMNS])
    for col in range(1, len(_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append(row.as_cells())
        written = ws.max_row
        for col, (_header, _name, kind) in enumerate(_COLUMNS, start=1):
            if kind == _MONEY:
                ws.cell(row=written, column=col).number_format = "#,##0"
            elif kind == _DAYS:
                ws.cell(row=written, column=col).number_format = "0.00"

    ws.freeze_panes = "A2"
    _autofit(ws, rows)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
