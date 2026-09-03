"""Payroll — the salary calculator, the org estimate, and HR/Admin-only authz.

The calculator is tested as a pure function (it must reproduce the reference
₹50,000 slip exactly). The API tests prove the wiring and, per CLAUDE §9, that
no one outside HR/Admin can reach org-wide payroll.
"""

from __future__ import annotations

import re
import uuid
from datetime import date, timedelta
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import Settings
from app.core.payroll import (
    CalcConfig,
    compute_breakdown,
    days_in_month,
    employer_contributions,
    payable_base_days,
    prorate_breakdown,
    weekdays_in_month,
    working_days_between,
)
from app.models.employee import Employee, EmployeeStatus, Role
from app.services.payroll_service import PayrollService
from tests.conftest import _Seed, auth_headers

# ---- pure calculator (the reference ₹50,000 structure, in paise) ----------- #


def test_working_days_per_week_counts_from_monday() -> None:
    # Any 7 consecutive days contain each weekday exactly once, so the count in a
    # full week equals the configured working-days-per-week (weekday-agnostic).
    start = date(2026, 6, 1)
    week = (start, start + timedelta(days=6))
    assert working_days_between(*week, set()) == 5  # default (Mon-Fri)
    assert working_days_between(*week, set(), 6) == 6  # Mon-Sat
    assert working_days_between(*week, set(), 7) == 7  # every day
    # June 2026 has 30 days; wdpw=7 counts them all, and Mon-Sat ≥ Mon-Fri.
    assert len(weekdays_in_month(2026, 6, 7)) == 30
    assert len(weekdays_in_month(2026, 6, 6)) >= len(weekdays_in_month(2026, 6, 5))


def test_breakdown_matches_reference_slip() -> None:
    b = compute_breakdown(50_000_00)  # ₹50,000/mo CTC
    assert b.basic_minor == 15_000_00
    assert b.hra_minor == 7_500_00
    assert b.special_allowance_minor == 25_700_00
    assert b.gross_minor == 48_200_00
    assert b.employer_pf_minor == 1_800_00
    assert b.employee_pf_minor == 1_800_00
    assert b.professional_tax_minor == 200_00
    assert b.total_deduction_minor == 2_000_00
    assert b.net_minor == 46_200_00


def test_breakdown_zero_ctc_is_all_zero() -> None:
    b = compute_breakdown(0)
    assert b.net_minor == 0 and b.gross_minor == 0 and b.basic_minor == 0


def test_breakdown_honours_config_knobs() -> None:
    b = compute_breakdown(50_000_00, CalcConfig(basic_pct=40, professional_tax_minor=0))
    assert b.basic_minor == 20_000_00
    assert b.professional_tax_minor == 0


def test_prorate_breakdown_matches_reference_sheet() -> None:
    # The reference "Salary Breakdown" sheet: ₹40,000/mo CTC, 26 of 30 days paid.
    b = compute_breakdown(40_000_00)
    assert b.gross_minor == 38_560_00  # full-month gross
    assert b.net_minor == 36_920_00  # full-month net (no TDS at this income)

    p = prorate_breakdown(b, payable_days=26, total_days=30)
    # Earnings + both PF figures scale by 26/30.
    assert p.basic_minor == 10_400_00
    assert p.hra_minor == 5_200_00
    assert p.special_allowance_minor == 17_818_67  # 20,560 * 26/30, to the paise
    assert p.gross_minor == 33_418_67  # displays as ₹33,419 (rounded)
    assert p.employee_pf_minor == 1_248_00
    # Professional tax stays FLAT — it is not attendance-linked.
    assert p.professional_tax_minor == 200_00
    assert p.income_tax_minor == 0
    assert p.total_deduction_minor == 1_448_00
    assert p.net_minor == 31_970_67  # displays as ₹31,971 (rounded)


def test_prorate_breakdown_edges() -> None:
    b = compute_breakdown(40_000_00)
    assert prorate_breakdown(b, 30, 30).net_minor == b.net_minor  # full month
    assert prorate_breakdown(b, 0, 30).gross_minor == 0  # nothing payable
    assert prorate_breakdown(b, 40, 30).net_minor == b.net_minor  # capped at full
    assert prorate_breakdown(b, 10, 0).net_minor == b.net_minor  # no days -> full
    assert days_in_month(2026, 2) == 28 and days_in_month(2026, 6) == 30


# ---- estimate over the org ------------------------------------------------- #

_COMP = {"amount_minor": 50_000_00, "currency": "inr", "period": "monthly"}
_SETTINGS = {
    "pay_day_of_month": 1,
    "currency": "INR",
    "pay_cycle": "monthly",
    "auto_send_enabled": False,
    "recipients": [],
    "basic_pct": 30,
    "hra_pct": 50,
    "pf_pct": 12,
    "pf_cap_minor": 1_800_00,
    "professional_tax_minor": 200_00,
    "professional_tax_feb_minor": 300_00,
    "deduct_income_tax": True,
}


async def test_estimate_computes_slip_and_total(
    client: AsyncClient, settings: Settings, seed: _Seed
) -> None:
    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    await client.put(
        "/api/v1/payroll/settings", json=_SETTINGS, headers=auth_headers(settings, seed.admin)
    )

    resp = await client.get(
        "/api/v1/payroll/estimate?month=2026-06", headers=auth_headers(settings, seed.admin)
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["currency"] == "INR"
    assert body["total_days"] == 30  # June 2026 is a 30-day month
    assert body["working_days"] > 0
    assert body["employee_count"] == len(body["lines"])
    assert body["total_net_minor"] == sum(line["net_minor"] for line in body["lines"])

    report_line = next(line for line in body["lines"] if line["employee_id"] == str(seed.report.id))
    # The full-month slip is independent of attendance and must match the image.
    assert report_line["monthly_ctc_minor"] == 50_000_00
    assert report_line["breakdown"]["net_minor"] == 46_200_00
    assert report_line["missing_compensation"] is False
    # No work sessions seeded -> every working day is loss-of-pay, but weekends and
    # holidays are still paid, so the prorated take-home is a fraction, not zero.
    assert report_line["net_minor"] == report_line["prorated"]["net_minor"]
    assert 0 < report_line["net_minor"] < 46_200_00
    assert report_line["payable_days"] == 30 - report_line["working_days"]

    outsider_line = next(
        line for line in body["lines"] if line["employee_id"] == str(seed.outsider.id)
    )
    assert outsider_line["missing_compensation"] is True
    assert outsider_line["monthly_ctc_minor"] == 0
    assert outsider_line["breakdown"]["net_minor"] == 0


async def test_future_working_days_are_not_charged_as_lop(
    client: AsyncClient, settings: Settings, seed: _Seed
) -> None:
    """A working day that has not happened yet is never loss-of-pay. For a month
    entirely in the future, an employee with zero attendance incurs zero LOP and is
    paid the full month — no more counting the calendar ahead as 'absent'."""
    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    await client.put(
        "/api/v1/payroll/settings", json=_SETTINGS, headers=auth_headers(settings, seed.admin)
    )

    resp = await client.get(
        "/api/v1/payroll/estimate?month=2099-06", headers=auth_headers(settings, seed.admin)
    )
    assert resp.status_code == 200
    line = next(line for line in resp.json()["lines"] if line["employee_id"] == str(seed.report.id))
    # The month has working days, but none have elapsed, so none are loss-of-pay.
    assert line["working_days"] > 0
    assert line["elapsed_working_days"] == 0
    assert line["lop_days"] == 0
    assert line["payable_days"] == line["total_days"]
    assert line["net_minor"] == line["breakdown"]["net_minor"]  # full month pay


async def test_payroll_scheduler_build_service_matches_constructor(db: AsyncSession) -> None:
    """The payroll worker wires PayrollService by hand (not FastAPI DI) and is not
    covered by `mypy app`, so a constructor change can silently break the running
    scheduler (it did: a missing `settings` arg took the worker down). Building the
    service here fails loudly if the wiring and the constructor drift apart."""
    from worker.payroll_scheduler import _build_service

    service = _build_service(db)
    assert isinstance(service, PayrollService)


async def test_hr_can_read_estimate(
    client: AsyncClient, settings: Settings, seed: _Seed, db: AsyncSession
) -> None:
    hr = Employee(
        hr_external_id="hr-person",
        work_email="hr@corp.test",
        full_name="Hank HR",
        role=Role.HR,
        status=EmployeeStatus.ACTIVE,
        is_active=True,
    )
    db.add(hr)
    await db.commit()

    resp = await client.get(
        "/api/v1/payroll/estimate?month=2026-06", headers=auth_headers(settings, hr)
    )
    assert resp.status_code == 200


# ---- excel export ---------------------------------------------------------- #

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def test_export_xlsx_hr_admin_only(
    client: AsyncClient, settings: Settings, seed: _Seed
) -> None:
    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    ok = await client.get(
        "/api/v1/payroll/export?month=2026-06", headers=auth_headers(settings, seed.admin)
    )
    assert ok.status_code == 200
    assert ok.headers["content-type"] == _XLSX_MEDIA_TYPE
    assert ok.content[:2] == b"PK"  # xlsx is a zip archive
    assert "payroll-2026-06.xlsx" in ok.headers["content-disposition"]

    for actor in (seed.report, seed.manager, seed.outsider):
        forbidden = await client.get(
            "/api/v1/payroll/export?month=2026-06", headers=auth_headers(settings, actor)
        )
        assert forbidden.status_code == 403, f"{actor.work_email} must not export payroll"


async def test_export_xlsx_selected_employees_only(
    client: AsyncClient, settings: Settings, seed: _Seed
) -> None:
    """`employee_ids` narrows the register to a chosen few, and the filename says so."""
    for emp in (seed.report, seed.manager):
        await client.put(
            f"/api/v1/employees/{emp.id}/compensation",
            json=_COMP,
            headers=auth_headers(settings, seed.admin),
        )

    everyone = await client.get(
        "/api/v1/payroll/export?month=2026-06", headers=auth_headers(settings, seed.admin)
    )
    assert everyone.status_code == 200
    assert "payroll-2026-06.xlsx" in everyone.headers["content-disposition"]

    one = await client.get(
        f"/api/v1/payroll/export?month=2026-06&employee_ids={seed.report.id}",
        headers=auth_headers(settings, seed.admin),
    )
    assert one.status_code == 200
    assert one.content[:2] == b"PK"
    # Named after the person, so a one-row register can't be mistaken for the org's.
    slug = re.sub(r"[^a-z0-9]+", "-", seed.report.full_name.lower()).strip("-")
    assert f"payroll-2026-06-{slug}.xlsx" in one.headers["content-disposition"]
    assert _sheet_row_by_name(one.content, seed.report.full_name)
    with pytest.raises(AssertionError):
        _sheet_row_by_name(one.content, seed.manager.full_name)

    two = await client.get(
        f"/api/v1/payroll/export?month=2026-06"
        f"&employee_ids={seed.report.id}&employee_ids={seed.manager.id}",
        headers=auth_headers(settings, seed.admin),
    )
    assert two.status_code == 200
    assert "payroll-2026-06-2-employees.xlsx" in two.headers["content-disposition"]


async def test_export_xlsx_selection_cannot_widen_scope(
    client: AsyncClient, settings: Settings, seed: _Seed
) -> None:
    """Selecting someone is a filter, not a grant: a non-HR caller is still 403, and
    an id outside the month's lines yields nothing rather than leaking a row."""
    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    forbidden = await client.get(
        f"/api/v1/payroll/export?month=2026-06&employee_ids={seed.report.id}",
        headers=auth_headers(settings, seed.manager),
    )
    assert forbidden.status_code == 403

    unknown = await client.get(
        f"/api/v1/payroll/export?month=2026-06&employee_ids={uuid.uuid4()}",
        headers=auth_headers(settings, seed.admin),
    )
    assert unknown.status_code == 404


def _sheet_row_by_name(content: bytes, name: str) -> dict[str, object]:
    """Load the register and return the {header: value} row for one employee."""
    ws = load_workbook(BytesIO(content)).active
    assert ws is not None
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = dict(zip(headers, row, strict=False))
        if cells.get("Employee Name") == name:
            return cells
    raise AssertionError(f"{name} not in export")


def test_employer_contributions_reference_split() -> None:
    # Basic ₹15,000 (the ₹50k structure): the employer 12% splits per the register.
    er = employer_contributions(15_000_00, 1_800_00)
    assert er.eps_minor == 1_250_00  # 8.33% of ₹15k ceiling, capped at ₹1,250
    assert er.epf_employer_minor == 550_00  # employee PF (1800) minus EPS (1250)
    assert er.edli_minor == 75_00  # 0.5% of ₹15k
    assert er.admin_minor == 75_00  # 0.5% of Basic
    assert er.total_minor == 1_950_00
    # A ₹9,870 basic (below the ceiling): everything scales off actual basic.
    low = employer_contributions(9_870_00, 1_184_00)
    assert low.eps_minor == round(9_870_00 * 8.33 / 100)
    assert low.epf_employer_minor == 1_184_00 - low.eps_minor


async def test_export_register_columns_and_consistency(
    client: AsyncClient, settings: Settings, seed: _Seed
) -> None:
    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    res = await client.get(
        "/api/v1/payroll/export?month=2026-06", headers=auth_headers(settings, seed.admin)
    )
    ws = load_workbook(BytesIO(res.content)).active
    assert ws is not None
    headers = [c.value for c in ws[1]]
    assert len(headers) == 40
    assert headers[0] == "Period" and headers[-1] == "Business Expense Reimbursements"

    row = _sheet_row_by_name(res.content, "Remy Report")
    assert row["Period"] == "June 2026"
    assert row["Payroll Type"] == "Regular Payroll"
    assert row["CTC Amount(Per Annum)"] == 600000  # 50,000 x 12
    assert row["Base Days"] == 30
    # The register is internally consistent regardless of how much LOP applies.
    assert row["Effective Paid Days"] == 30 - row["Loss Of Pay"]
    assert (
        row["EPF Contribution Employer"]
        == row["EPF Contribution"] - row["EPS Contribution Employer"]
    )
    assert row["Total Employer Contributions"] == (
        row["EPF Contribution Employer"]
        + row["EPS Contribution Employer"]
        + row["Employer EDLI Contribution Employer"]
        + row["Employer EPF Admin Charges Employer"]
    )
    assert row["Total Deductions"] == (
        row["EPF Contribution"] + row["Income Tax"] + row["Maharashtra Professional Tax"]
    )
    assert row["Net Pay"] == row["Gross Pay"] - row["Total Deductions"]
    assert row["Business Expense Reimbursements"] == 0  # none approved yet


async def test_live_payslip_pdf_self_or_hr_admin(
    client: AsyncClient, settings: Settings, seed: _Seed
) -> None:
    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    url = "/api/v1/payroll/me/pdf?month=2026-06"

    # HR/Admin generate anyone's live PDF before release (a preview).
    theirs = await client.get(
        f"{url}&employee_id={seed.report.id}", headers=auth_headers(settings, seed.admin)
    )
    assert theirs.status_code == 200
    assert theirs.content[:4] == b"%PDF"

    # The individual can't generate their own until the month is released.
    early = await client.get(url, headers=auth_headers(settings, seed.report))
    assert early.status_code == 404
    await client.post(
        "/api/v1/payroll/finalize?month=2026-06", headers=auth_headers(settings, seed.admin)
    )
    mine = await client.get(url, headers=auth_headers(settings, seed.report))
    assert mine.status_code == 200, mine.text
    assert mine.content[:4] == b"%PDF"
    assert "payslip-2026-06.pdf" in mine.headers["content-disposition"]

    # A manager or unrelated employee cannot generate someone else's (need-to-know).
    for actor in (seed.manager, seed.outsider):
        denied = await client.get(
            f"{url}&employee_id={seed.report.id}", headers=auth_headers(settings, actor)
        )
        assert denied.status_code == 403, f"{actor.work_email} must not generate another's payslip"


async def test_export_includes_approved_reimbursement(
    client: AsyncClient, settings: Settings, seed: _Seed, db: AsyncSession
) -> None:
    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    # Report submits a ₹500 claim for June; manager then HR approve it.
    claim = await client.post(
        "/api/v1/reimbursements",
        json={
            "amount_minor": 500_00,
            "category": "travel",
            "description": "June cab",
            "expense_date": "2026-06-10",
        },
        headers=auth_headers(settings, seed.report),
    )
    assert claim.status_code == 201, claim.text
    rid = claim.json()["id"]
    await client.post(
        f"/api/v1/reimbursements/{rid}/manager-decision",
        json={"approve": True},
        headers=auth_headers(settings, seed.manager),
    )
    # HR, not Admin: reimbursement approval is HR's (and payroll-grant holders'),
    # since a claim is personal spending data.
    reviewer = Employee(
        hr_external_id="hr-export-reviewer",
        work_email="export-reviewer@corp.test",
        full_name="Hana HR",
        role=Role.HR,
        status=EmployeeStatus.ACTIVE,
        is_active=True,
    )
    db.add(reviewer)
    await db.commit()
    approved = await client.post(
        f"/api/v1/reimbursements/{rid}/hr-decision",
        json={"approve": True},
        headers=auth_headers(settings, reviewer),
    )
    assert approved.status_code == 200, approved.text

    res = await client.get(
        "/api/v1/payroll/export?month=2026-06", headers=auth_headers(settings, seed.admin)
    )
    row = _sheet_row_by_name(res.content, "Remy Report")
    assert row["Business Expense Reimbursements"] == 500


# ---- authorization: payroll is HR/Admin only (CLAUDE §9) ------------------- #


async def test_estimate_forbidden_for_non_hr_admin(
    client: AsyncClient, settings: Settings, seed: _Seed
) -> None:
    for actor in (seed.report, seed.manager, seed.outsider):
        resp = await client.get(
            "/api/v1/payroll/estimate?month=2026-06", headers=auth_headers(settings, actor)
        )
        assert resp.status_code == 403, f"{actor.work_email} should not read payroll"


async def test_settings_and_send_forbidden_for_non_hr_admin(
    client: AsyncClient, settings: Settings, seed: _Seed
) -> None:
    get_settings = await client.get(
        "/api/v1/payroll/settings", headers=auth_headers(settings, seed.manager)
    )
    assert get_settings.status_code == 403

    put_settings = await client.put(
        "/api/v1/payroll/settings", json=_SETTINGS, headers=auth_headers(settings, seed.manager)
    )
    assert put_settings.status_code == 403

    send = await client.post(
        "/api/v1/payroll/send?month=2026-06", headers=auth_headers(settings, seed.report)
    )
    assert send.status_code == 403

    runs = await client.get("/api/v1/payroll/runs", headers=auth_headers(settings, seed.outsider))
    assert runs.status_code == 403


# ---- send digest ----------------------------------------------------------- #


async def test_send_requires_recipients_then_records_run(
    client: AsyncClient, settings: Settings, seed: _Seed
) -> None:
    # No recipients configured yet -> a helpful 422, not a silent no-op.
    empty = await client.post(
        "/api/v1/payroll/send?month=2026-06", headers=auth_headers(settings, seed.admin)
    )
    assert empty.status_code == 422

    await client.put(
        "/api/v1/payroll/settings",
        json={**_SETTINGS, "recipients": ["finance@corp.test", "hr@corp.test"]},
        headers=auth_headers(settings, seed.admin),
    )
    sent = await client.post(
        "/api/v1/payroll/send?month=2026-06", headers=auth_headers(settings, seed.admin)
    )
    assert sent.status_code == 200
    run = sent.json()
    assert run["period_month"] == "2026-06"
    assert run["source"] == "manual"
    assert run["recipients"] == ["finance@corp.test", "hr@corp.test"]

    runs = await client.get("/api/v1/payroll/runs", headers=auth_headers(settings, seed.admin))
    assert runs.status_code == 200
    assert any(r["period_month"] == "2026-06" for r in runs.json())


async def test_approved_reimbursement_lands_in_the_payroll_breakdown(
    client: AsyncClient, settings: Settings, seed: _Seed, db: AsyncSession
) -> None:
    """An approved claim is paid with that month's salary, so it must be visible
    on the line — not only in the register export.

    It is added AFTER PF and tax and stays out of `prorated`: repaying an expense
    is not earnings and must never inflate gross or be taxed.
    """
    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    month = date.today().strftime("%Y-%m")

    before = await client.get(
        f"/api/v1/payroll/estimate?month={month}", headers=auth_headers(settings, seed.admin)
    )
    assert before.status_code == 200, before.text
    base = next(r for r in before.json()["lines"] if r["employee_id"] == str(seed.report.id))
    assert base["reimbursement_minor"] == 0

    claim = await client.post(
        "/api/v1/reimbursements",
        json={
            "amount_minor": 500_00,
            "category": "travel",
            "description": "Client cab",
            "expense_date": date.today().isoformat(),
        },
        headers=auth_headers(settings, seed.report),
    )
    assert claim.status_code == 201, claim.text
    rid = claim.json()["id"]
    await client.post(
        f"/api/v1/reimbursements/{rid}/manager-decision",
        json={"approve": True},
        headers=auth_headers(settings, seed.manager),
    )
    reviewer = Employee(
        hr_external_id="hr-breakdown-reviewer",
        work_email="breakdown-reviewer@corp.test",
        full_name="Hana HR",
        role=Role.HR,
        status=EmployeeStatus.ACTIVE,
        is_active=True,
    )
    db.add(reviewer)
    await db.commit()
    approved = await client.post(
        f"/api/v1/reimbursements/{rid}/hr-decision",
        json={"approve": True},
        headers=auth_headers(settings, reviewer),
    )
    assert approved.status_code == 200, approved.text

    after = await client.get(
        f"/api/v1/payroll/estimate?month={month}", headers=auth_headers(settings, seed.admin)
    )
    line = next(r for r in after.json()["lines"] if r["employee_id"] == str(seed.report.id))
    assert line["reimbursement_minor"] == 500_00
    assert line["net_minor"] == base["net_minor"] + 500_00
    # Untouched: taxes and PF are computed on salary alone.
    assert line["prorated"]["gross_minor"] == base["prorated"]["gross_minor"]
    assert line["prorated"]["employee_pf_minor"] == base["prorated"]["employee_pf_minor"]


# --- month of joining -------------------------------------------------------- #
def test_payroll_window_starts_at_the_hire_date_only_in_the_joining_month() -> None:
    from app.core.payroll import payroll_window

    # Joined mid-August: on the payroll for 15-31 Aug = 17 of 31 days.
    assert payroll_window(2026, 8, date(2026, 8, 15)) == (date(2026, 8, 15), date(2026, 8, 31))
    assert payable_base_days(2026, 8, date(2026, 8, 15)) == 17

    # Every later month is whole, and so is a month before nothing changed.
    assert payable_base_days(2026, 9, date(2026, 8, 15)) == 30
    assert payable_base_days(2026, 8, None) == 31
    assert payable_base_days(2026, 8, date(2020, 1, 1)) == 31


def test_payable_base_days_handles_every_month_length() -> None:
    """30/31, and February in both a leap and a common year."""
    assert payable_base_days(2026, 2, None) == 28  # common year
    assert payable_base_days(2028, 2, None) == 29  # leap year
    assert payable_base_days(2026, 4, None) == 30
    assert payable_base_days(2026, 1, None) == 31
    # Joining on the last day earns exactly one day, in any month length.
    assert payable_base_days(2026, 2, date(2026, 2, 28)) == 1
    assert payable_base_days(2026, 4, date(2026, 4, 30)) == 1
    # Joining on the 1st is a whole month, not a day short.
    assert payable_base_days(2026, 6, date(2026, 6, 1)) == 30
    # Not joined yet this month → nothing is payable.
    assert payable_base_days(2026, 5, date(2026, 7, 1)) == 0


async def test_a_mid_month_joiner_is_not_paid_for_the_days_before_they_joined(
    client: AsyncClient, db: AsyncSession, settings: Settings, seed: _Seed
) -> None:
    """The bug: LOP only ever lands on WORKING days, so the weekends sitting before
    a mid-month start date were silently paid — a 15 Aug joiner collected the 1st,
    2nd, 8th and 9th. The payable base now starts at the hire date."""
    # A future month, so nothing has elapsed and no real absence can muddy the
    # arithmetic — this isolates the joining window itself.
    person = await db.get(Employee, seed.report.id)
    assert person is not None
    person.hire_date = date(2099, 6, 15)
    await db.commit()

    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    await client.put(
        "/api/v1/payroll/settings", json=_SETTINGS, headers=auth_headers(settings, seed.admin)
    )

    resp = await client.get(
        "/api/v1/payroll/estimate?month=2099-06", headers=auth_headers(settings, seed.admin)
    )
    assert resp.status_code == 200, resp.text
    line = next(line for line in resp.json()["lines"] if line["employee_id"] == str(seed.report.id))

    assert line["total_days"] == 30  # June is still a 30-day month...
    assert line["payable_days"] == 16  # ...but only 15-30 June is theirs
    assert line["lop_days"] == 0  # and nothing before the 15th is absence

    # Paid 16/30 of the month, not the whole thing — the old behaviour paid all 30,
    # because the weekends before the 15th were never eligible for loss-of-pay.
    full = line["breakdown"]["gross_minor"]
    assert line["prorated"]["gross_minor"] == round(full * 16 / 30)
    assert line["prorated"]["gross_minor"] < full

    # Working days are scoped to their window too, so the slip reads truthfully.
    assert line["working_days"] == working_days_between(
        date(2099, 6, 15), date(2099, 6, 30), set(), 5
    )


async def test_someone_already_on_the_roster_is_unaffected(
    client: AsyncClient, db: AsyncSession, settings: Settings, seed: _Seed
) -> None:
    """The joining-month rule must not touch anyone else's slip."""
    person = await db.get(Employee, seed.report.id)
    assert person is not None
    person.hire_date = date(2020, 1, 6)
    await db.commit()

    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    await client.put(
        "/api/v1/payroll/settings", json=_SETTINGS, headers=auth_headers(settings, seed.admin)
    )

    resp = await client.get(
        "/api/v1/payroll/estimate?month=2099-06", headers=auth_headers(settings, seed.admin)
    )
    line = next(line for line in resp.json()["lines"] if line["employee_id"] == str(seed.report.id))
    assert line["payable_days"] == line["total_days"] == 30
    assert line["net_minor"] == line["breakdown"]["net_minor"]


async def test_the_register_pays_a_joiner_a_part_month_too(
    client: AsyncClient, db: AsyncSession, settings: Settings, seed: _Seed
) -> None:
    """The Excel register is the file that actually gets paid, and it prorates on a
    fixed 30-day base of its own. That base has to shrink to the joining window too,
    or the screen shows a part month while the payrun hands over a full one."""
    person = await db.get(Employee, seed.report.id)
    assert person is not None
    person.hire_date = date(2099, 6, 15)  # 16 of June's 30 days
    await db.commit()

    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    await client.put(
        "/api/v1/payroll/settings", json=_SETTINGS, headers=auth_headers(settings, seed.admin)
    )

    resp = await client.get(
        "/api/v1/payroll/export?month=2099-06", headers=auth_headers(settings, seed.admin)
    )
    assert resp.status_code == 200, resp.text
    sheet = load_workbook(BytesIO(resp.content)).active
    assert sheet is not None
    header = [c.value for c in sheet[1]]
    rows = {
        r[header.index("Employee Name")]: r for r in sheet.iter_rows(min_row=2, values_only=True)
    }
    row = rows[seed.report.full_name]

    base = row[header.index("Base Days")]
    paid = row[header.index("Effective Paid Days")]
    # 30 x 16/30 = 16 base days, none lost — a part month, not 30.
    assert base == pytest.approx(16.0)
    assert paid == pytest.approx(16.0)
    assert row[header.index("Loss Of Pay")] == pytest.approx(0.0)

    # The day columns being right is not the point — the MONEY has to be right.
    # (A denominator slip here once produced a ratio above 1, which clamped back
    # to a full month while these day columns still read 16.)
    full_month_gross = row[header.index("Gross Amount(Per Annum)")] / 12
    assert row[header.index("Gross Pay")] == pytest.approx(full_month_gross * 16 / 30, rel=0.01)
    assert row[header.index("Gross Pay")] < full_month_gross


async def test_the_register_still_gives_everyone_else_a_full_30_day_base(
    client: AsyncClient, db: AsyncSession, settings: Settings, seed: _Seed
) -> None:
    """The joining-month rule must not disturb the register's own convention."""
    person = await db.get(Employee, seed.report.id)
    assert person is not None
    person.hire_date = date(2020, 1, 6)
    await db.commit()

    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    await client.put(
        "/api/v1/payroll/settings", json=_SETTINGS, headers=auth_headers(settings, seed.admin)
    )

    resp = await client.get(
        "/api/v1/payroll/export?month=2099-06", headers=auth_headers(settings, seed.admin)
    )
    sheet = load_workbook(BytesIO(resp.content)).active
    assert sheet is not None
    header = [c.value for c in sheet[1]]
    rows = {
        r[header.index("Employee Name")]: r for r in sheet.iter_rows(min_row=2, values_only=True)
    }
    row = rows[seed.report.full_name]
    assert row[header.index("Base Days")] == pytest.approx(30.0)
    # And a whole month's money, not a fraction of one.
    assert row[header.index("Gross Pay")] == pytest.approx(
        row[header.index("Gross Amount(Per Annum)")] / 12, rel=0.01
    )


async def test_someone_who_has_not_joined_yet_is_paid_nothing_anywhere(
    client: AsyncClient, db: AsyncSession, settings: Settings, seed: _Seed
) -> None:
    """A hire date after the month means a zero-day window. `attendance_ratio`
    reads a zero denominator as "no information — pay the full month", which is
    the exact opposite of what it means here, so both the estimate and the
    register must be checked, not just one."""
    person = await db.get(Employee, seed.report.id)
    assert person is not None
    person.hire_date = date(2099, 9, 1)  # joins in September
    await db.commit()

    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    await client.put(
        "/api/v1/payroll/settings", json=_SETTINGS, headers=auth_headers(settings, seed.admin)
    )

    est = await client.get(
        "/api/v1/payroll/estimate?month=2099-06", headers=auth_headers(settings, seed.admin)
    )
    line = next(r for r in est.json()["lines"] if r["employee_id"] == str(seed.report.id))
    assert line["payable_base_days"] == 0
    assert line["payable_days"] == 0
    assert line["prorated"]["gross_minor"] == 0

    export = await client.get(
        "/api/v1/payroll/export?month=2099-06", headers=auth_headers(settings, seed.admin)
    )
    sheet = load_workbook(BytesIO(export.content)).active
    assert sheet is not None
    header = [c.value for c in sheet[1]]
    rows = {
        r[header.index("Employee Name")]: r for r in sheet.iter_rows(min_row=2, values_only=True)
    }
    row = rows[seed.report.full_name]
    assert row[header.index("Base Days")] == pytest.approx(0.0)
    assert row[header.index("Gross Pay")] == pytest.approx(0.0)


async def _override_payable_days(
    client: AsyncClient, settings: Settings, seed: _Seed, *, month: str, days: float
) -> None:
    resp = await client.post(
        "/api/v1/payroll/adjustments",
        json={
            "employee_id": str(seed.report.id),
            "period_month": month,
            "kind": "override",
            "target": "payable_days",
            "label": "Agreed paid days",
            "amount_minor": int(days * 100),
        },
        headers=auth_headers(settings, seed.admin),
    )
    assert resp.status_code in (200, 201), resp.text


async def test_editing_payable_days_moves_the_money_on_screen_and_in_the_register(
    client: AsyncClient, settings: Settings, seed: _Seed
) -> None:
    """Changing payable days must restate the pay everywhere it is shown, not just
    the day counter — the estimate, the prorated breakdown, the net, and the Excel
    register that is actually paid from."""
    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    await client.put(
        "/api/v1/payroll/settings", json=_SETTINGS, headers=auth_headers(settings, seed.admin)
    )
    await _override_payable_days(client, settings, seed, month="2099-06", days=15.0)

    est = await client.get(
        "/api/v1/payroll/estimate?month=2099-06", headers=auth_headers(settings, seed.admin)
    )
    line = next(r for r in est.json()["lines"] if r["employee_id"] == str(seed.report.id))
    assert line["payable_days"] == 15.0
    assert line["lop_days"] == 15.0  # base 30 - 15 paid
    # Half a month of days must be half a month of money.
    full = line["breakdown"]["gross_minor"]
    assert line["prorated"]["gross_minor"] == round(full * 15 / 30)
    assert line["net_minor"] < line["breakdown"]["net_minor"]

    export = await client.get(
        "/api/v1/payroll/export?month=2099-06", headers=auth_headers(settings, seed.admin)
    )
    sheet = load_workbook(BytesIO(export.content)).active
    assert sheet is not None
    header = [c.value for c in sheet[1]]
    rows = {
        r[header.index("Employee Name")]: r for r in sheet.iter_rows(min_row=2, values_only=True)
    }
    row = rows[seed.report.full_name]
    assert row[header.index("Effective Paid Days")] == pytest.approx(15.0)
    assert row[header.index("Gross Pay")] == pytest.approx(
        row[header.index("Gross Amount(Per Annum)")] / 12 * 15 / 30, rel=0.01
    )


async def test_editing_payable_days_for_a_joiner_stays_inside_their_window(
    client: AsyncClient, db: AsyncSession, settings: Settings, seed: _Seed
) -> None:
    """For a mid-month joiner the override is read against THEIR window, not the
    month: 10 paid days out of 16 is 6 unpaid, not 20."""
    person = await db.get(Employee, seed.report.id)
    assert person is not None
    person.hire_date = date(2099, 6, 15)  # 16 of June's 30 days
    await db.commit()

    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    await client.put(
        "/api/v1/payroll/settings", json=_SETTINGS, headers=auth_headers(settings, seed.admin)
    )
    await _override_payable_days(client, settings, seed, month="2099-06", days=10.0)

    est = await client.get(
        "/api/v1/payroll/estimate?month=2099-06", headers=auth_headers(settings, seed.admin)
    )
    line = next(r for r in est.json()["lines"] if r["employee_id"] == str(seed.report.id))
    assert line["payable_base_days"] == 16
    assert line["payable_days"] == 10.0
    assert line["lop_days"] == 6.0  # 16 - 10, NOT 30 - 10
    full = line["breakdown"]["gross_minor"]
    assert line["prorated"]["gross_minor"] == round(full * 10 / 30)


async def test_a_payable_days_override_above_the_window_is_capped_not_negative(
    client: AsyncClient, db: AsyncSession, settings: Settings, seed: _Seed
) -> None:
    """Asking for more paid days than exist can only ever mean "all of them" —
    never negative loss-of-pay, which would quietly pay more than a full month."""
    person = await db.get(Employee, seed.report.id)
    assert person is not None
    person.hire_date = date(2099, 6, 15)
    await db.commit()

    await client.put(
        f"/api/v1/employees/{seed.report.id}/compensation",
        json=_COMP,
        headers=auth_headers(settings, seed.admin),
    )
    await _override_payable_days(client, settings, seed, month="2099-06", days=99.0)

    est = await client.get(
        "/api/v1/payroll/estimate?month=2099-06", headers=auth_headers(settings, seed.admin)
    )
    line = next(r for r in est.json()["lines"] if r["employee_id"] == str(seed.report.id))
    assert line["lop_days"] == 0.0
    assert line["payable_days"] == 16.0  # capped at their window
    full = line["breakdown"]["gross_minor"]
    assert line["prorated"]["gross_minor"] == round(full * 16 / 30)
